from http.server import BaseHTTPRequestHandler
import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.request
import urllib.parse

SUPABASE_URL = "https://pvwcpukfhyrmdpxgfwrk.supabase.co"
SUPABASE_ANON = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InB2d2NwdWtmaHlybWRweGdmd3JrIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY4NzIzNzMsImV4cCI6MjA5MjQ0ODM3M30.69GpTOQdwtAgouGN0MIHqQc6bImxAbfQ72xL2S4KaoM"

def fetch_debt():
    url = f"{SUPABASE_URL}/rest/v1/debt?order=days_overdue.desc"
    req = urllib.request.Request(url, headers={
        "apikey": SUPABASE_ANON,
        "Authorization": f"Bearer {SUPABASE_ANON}",
    })
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def S(bc="D1D5DB"):
    s = Side(style='thin', color=bc)
    return Border(left=s, right=s, top=s, bottom=s)

def make_excel(debt):
    overdue = sorted([d for d in debt if d['status']=='באיחור'], key=lambda x: -x['days_overdue'])
    current = sorted([d for d in debt if d['status']!='באיחור'], key=lambda x: -x['amount'])
    all_debt = overdue + current
    total = sum(d['amount'] for d in all_debt)
    total_ov = sum(d['amount'] for d in overdue)
    total_cu = sum(d['amount'] for d in current)

    wb = Workbook()

    def hdr(c, bg, fg="FFFFFF", sz=10):
        c.font = Font(name='Arial', bold=True, size=sz, color=fg)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = S(bg)

    def dat(c, bg, fg, bold=False, al='center'):
        c.font = Font(name='Arial', size=10, color=fg, bold=bold)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal=al, vertical='center')
        c.border = S()

    now = datetime.now().strftime('%d/%m/%Y')

    # ── Sheet 1: דשבורד ──
    ws = wb.active
    ws.title = "סיכום"
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "1E3A5F"

    for col, w in zip("ABCDEFG", [3, 34, 18, 16, 14, 16, 3]):
        ws.column_dimensions[col].width = w

    for r in range(1, 4):
        ws.row_dimensions[r].height = 14 if r in [1,3] else 32
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = PatternFill('solid', start_color="0F172A")

    ws.merge_cells('B2:F2')
    c = ws['B2']
    c.value = "דוח גבייה — בית דפוס"
    c.font = Font(name='Arial', bold=True, size=18, color="F8FAFC")
    c.fill = PatternFill('solid', start_color="0F172A")
    c.alignment = Alignment(horizontal='right', vertical='center')

    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 8

    for col, title, value, dark, light, mid in [
        ("B", "סה״כ חוב פתוח", f"₪{total:,.0f}", "1E40AF", "DBEAFE", "2563EB"),
        ("D", f"באיחור 🚨", f"₪{total_ov:,.0f}", "991B1B", "FEE2E2", "DC2626"),
        ("F", "שוטף ✅", f"₪{total_cu:,.0f}", "14532D", "DCFCE7", "16A34A"),
    ]:
        nc = chr(ord(col)+1)
        ws.merge_cells(f'{col}5:{nc}5')
        ws.merge_cells(f'{col}6:{nc}6')
        tc = ws[f'{col}5']
        tc.value = title
        tc.font = Font(name='Arial', bold=True, size=11, color=dark)
        tc.fill = PatternFill('solid', start_color=light)
        tc.alignment = Alignment(horizontal='center', vertical='bottom')
        tc.border = Border(top=Side(style='medium', color=mid), left=Side(style='thin', color="E2E8F0"), right=Side(style='thin', color="E2E8F0"))
        vc = ws[f'{col}6']
        vc.value = value
        vc.font = Font(name='Arial', bold=True, size=18, color=dark)
        vc.fill = PatternFill('solid', start_color=light)
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.border = Border(bottom=Side(style='medium', color=mid), left=Side(style='thin', color="E2E8F0"), right=Side(style='thin', color="E2E8F0"))

    ws.row_dimensions[8].height = 30
    ws.merge_cells('B8:F8')
    st = ws['B8']
    st.value = f"פירוט חייבים — {now}  |  {len(all_debt)} לקוחות"
    st.font = Font(name='Arial', bold=True, size=12, color="1E293B")
    st.alignment = Alignment(horizontal='right', vertical='center')
    st.border = Border(bottom=Side(style='medium', color="2563EB"))

    ws.row_dimensions[9].height = 28
    for h, ci in zip(["לקוח", "סכום חוב (₪)", "תאריך פירעון", "ימי איחור", "דחיפות"], [2,3,4,5,6]):
        hdr(ws.cell(row=9, column=ci, value=h), "1E3A5F")

    for ri, d in enumerate(all_debt, 10):
        ws.row_dimensions[ri].height = 22
        is_ov = d['status'] == 'באיחור'
        days = d['days_overdue']
        bg = ("FEF2F2" if ri%2==0 else "FFF7F7") if is_ov else ("F8FAFC" if ri%2==0 else "FFFFFF")
        urg = ("🔴 דחוף מאוד" if days>90 else "🟠 דחוף" if days>30 else "🟡 עקוב") if is_ov else "🟢 שוטף"
        uc = ("991B1B" if days>90 else "C2410C" if days>30 else "854D0E") if is_ov else "14532D"

        dat(ws.cell(row=ri, column=2, value=d['client_name']), bg, "1E293B", True, 'right')
        c = ws.cell(row=ri, column=3, value=d['amount'])
        dat(c, bg, "DC2626" if is_ov else "16A34A", True)
        c.number_format = '#,##0'
        dat(ws.cell(row=ri, column=4, value=d['due_date']), bg, "475569")
        dat(ws.cell(row=ri, column=5, value=days if is_ov else "—"), bg, "DC2626" if is_ov else "475569", is_ov)
        dat(ws.cell(row=ri, column=6, value=urg), bg, uc)

    tr = len(all_debt) + 10
    ws.row_dimensions[tr].height = 26
    for ci, val in enumerate(["סה״כ", total, "", "", ""], 2):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill('solid', start_color="1E3A5F")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = S("1E3A5F")
    ws.cell(row=tr, column=3).number_format = '#,##0'

    # ── Sheet 2: באיחור ──
    ws2 = wb.create_sheet("חייבים באיחור")
    ws2.sheet_view.rightToLeft = True
    ws2.sheet_properties.tabColor = "DC2626"

    for col, w in zip("ABCDEFG", [3, 34, 16, 16, 14, 20, 3]):
        ws2.column_dimensions[col].width = w

    for r in range(1, 4):
        ws2.row_dimensions[r].height = 14 if r in [1,3] else 38
        for c in range(1, 8):
            ws2.cell(row=r, column=c).fill = PatternFill('solid', start_color="450A0A")

    ws2.merge_cells('B2:F2')
    c = ws2['B2']
    c.value = f"חייבים באיחור — {len(overdue)} לקוחות | ₪{total_ov:,.0f}"
    c.font = Font(name='Arial', bold=True, size=14, color="FECACA")
    c.fill = PatternFill('solid', start_color="450A0A")
    c.alignment = Alignment(horizontal='right', vertical='center')

    ws2.row_dimensions[4].height = 28
    for h, ci in zip(["לקוח", "סכום (₪)", "תאריך פירעון", "ימי איחור", "דחיפות / הערות"], [2,3,4,5,6]):
        hdr(ws2.cell(row=4, column=ci, value=h), "DC2626")

    for ri, d in enumerate(overdue, 5):
        ws2.row_dimensions[ri].height = 24
        days = d['days_overdue']
        bg = "FEF2F2" if ri%2==0 else "FFFFFF"
        urg = "🔴 דחוף מאוד" if days>90 else "🟠 דחוף" if days>30 else "🟡 עקוב"
        uc = "991B1B" if days>90 else "C2410C" if days>30 else "854D0E"
        notes = d.get('payment_notes') or ""
        display = f"{urg}  {notes}" if notes else urg

        dat(ws2.cell(row=ri, column=2, value=d['client_name']), bg, "1E293B", True, 'right')
        c = ws2.cell(row=ri, column=3, value=d['amount'])
        dat(c, bg, "DC2626", True)
        c.number_format = '#,##0'
        dat(ws2.cell(row=ri, column=4, value=d['due_date']), bg, "475569")
        dat(ws2.cell(row=ri, column=5, value=days), bg, "DC2626", True)
        c2 = ws2.cell(row=ri, column=6, value=display)
        dat(c2, bg, uc, False, 'right')
        c2.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    tr2 = len(overdue) + 5
    for ci, val in enumerate(["סה״כ", total_ov, "", "", ""], 2):
        c = ws2.cell(row=tr2, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill('solid', start_color="DC2626")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = S("DC2626")
    ws2.cell(row=tr2, column=3).number_format = '#,##0'

    # ── Sheet 3: שוטף ──
    ws3 = wb.create_sheet("שוטף")
    ws3.sheet_view.rightToLeft = True
    ws3.sheet_properties.tabColor = "16A34A"

    for col, w in zip("ABCDE", [3, 34, 16, 16, 3]):
        ws3.column_dimensions[col].width = w

    for r in range(1, 4):
        ws3.row_dimensions[r].height = 14 if r in [1,3] else 38
        for c in range(1, 6):
            ws3.cell(row=r, column=c).fill = PatternFill('solid', start_color="052E16")

    ws3.merge_cells('B2:D2')
    c = ws3['B2']
    c.value = f"חוב שוטף — {len(current)} לקוחות | ₪{total_cu:,.0f}"
    c.font = Font(name='Arial', bold=True, size=14, color="BBF7D0")
    c.fill = PatternFill('solid', start_color="052E16")
    c.alignment = Alignment(horizontal='right', vertical='center')

    ws3.row_dimensions[4].height = 28
    for h, ci in zip(["לקוח", "סכום (₪)", "תאריך פירעון"], [2,3,4]):
        hdr(ws3.cell(row=4, column=ci, value=h), "16A34A")

    for ri, d in enumerate(current, 5):
        ws3.row_dimensions[ri].height = 22
        bg = "F0FDF4" if ri%2==0 else "FFFFFF"
        dat(ws3.cell(row=ri, column=2, value=d['client_name']), bg, "1E293B", True, 'right')
        c = ws3.cell(row=ri, column=3, value=d['amount'])
        dat(c, bg, "16A34A", True)
        c.number_format = '#,##0'
        dat(ws3.cell(row=ri, column=4, value=d['due_date']), bg, "475569")

    tr3 = len(current) + 5
    for ci, val in enumerate(["סה״כ", total_cu, ""], 2):
        c = ws3.cell(row=tr3, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill('solid', start_color="16A34A")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = S("16A34A")
    ws3.cell(row=tr3, column=3).number_format = '#,##0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            debt = fetch_debt()
            excel_bytes = make_excel(debt)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="דוח_גבייה.xlsx"')
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())
